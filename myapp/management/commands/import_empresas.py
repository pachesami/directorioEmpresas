"""
Importa empresas desde Excel con logos extraídos
"""
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from openpyxl import load_workbook
from myapp.models import Empresa
import shutil

class Command(BaseCommand):
    help = "Importa empresas desde Excel con logos en carpeta separada"

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Ruta al archivo Excel (.xlsx)')
        parser.add_argument(
            '--logos-dir',
            type=str,
            default='extracted_logos',
            help='Carpeta con los logos extraídos (default: extracted_logos)'
        )
        parser.add_argument(
            '--skip-logos',
            action='store_true',
            help='Importar sin logos'
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path'])
        logos_dir = Path(options['logos_dir'])
        skip_logos = options['skip_logos']

        # Validar archivo Excel
        if not xlsx_path.exists():
            raise CommandError(f"❌ No existe el archivo: {xlsx_path}")
        
        # Validar carpeta de logos
        logos_disponibles = []
        if not skip_logos:
            if not logos_dir.exists():
                self.stdout.write(self.style.WARNING(
                    f"⚠️  No existe la carpeta: {logos_dir}\n"
                    f"   Ejecuta primero: python extract_images_zip.py {xlsx_path.name}\n"
                    f"   O usa --skip-logos para importar sin imágenes"
                ))
                return
            
            logos_disponibles = sorted(logos_dir.glob('imagen_*.*'))
            if not logos_disponibles:
                self.stdout.write(self.style.WARNING(
                    f"⚠️  No hay logos en: {logos_dir}\n"
                    f"   Ejecuta: python extract_images_zip.py {xlsx_path.name}"
                ))
                return
            
            self.stdout.write(self.style.SUCCESS(f"📂 Logos encontrados: {len(logos_disponibles)}"))

        # Cargar Excel
        self.stdout.write(self.style.NOTICE(f"📂 Abriendo Excel: {xlsx_path.name}\n"))
        
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
        except Exception as e:
            raise CommandError(f"❌ Error abriendo Excel: {e}")

        # Crear carpeta de destino para logos
        media_logos = Path(settings.MEDIA_ROOT) / 'logos'
        media_logos.mkdir(parents=True, exist_ok=True)

        # Contadores
        created = 0
        updated = 0
        skipped = 0
        errors = 0

        # Procesar filas (asumiendo encabezados en fila 1)
        start_row = 2
        total_rows = ws.max_row

        self.stdout.write(self.style.NOTICE(f"📊 Procesando {total_rows - 1} filas...\n"))

        for idx, row_num in enumerate(range(start_row, total_rows + 1)):
            # Leer columnas: A=Cliente, B=Compañía, C=Código
            cliente_val = ws.cell(row=row_num, column=1).value
            compania_val = ws.cell(row=row_num, column=2).value
            codigo_val = ws.cell(row=row_num, column=3).value

            # Normalizar
            cliente = str(cliente_val).strip() if cliente_val else ""
            compania = str(compania_val).strip() if compania_val else ""
            codigo = str(codigo_val).strip() if codigo_val else ""

            # Saltar filas vacías
            if not (cliente or compania):
                skipped += 1
                continue

            self.stdout.write(f"Fila {row_num}: ", ending='')

            try:
                # Buscar empresa existente
                empresa = None
                if codigo:
                    empresa = Empresa.objects.filter(codigo=codigo).first()
                if not empresa and compania:
                    empresa = Empresa.objects.filter(compania__iexact=compania).first()

                # Asignar logo (la imagen en posición idx corresponde a esta fila)
                logo_field = None
                if not skip_logos and idx < len(logos_disponibles):
                    logo_path = logos_disponibles[idx]
                    
                    # Copiar logo a media/logos/
                    dest_filename = f"{codigo}_{logo_path.name}" if codigo else logo_path.name
                    dest_path = media_logos / dest_filename
                    
                    shutil.copy2(logo_path, dest_path)
                    logo_field = f"logos/{dest_filename}"
                    
                    self.stdout.write(f"📷 ", ending='')

                # Crear o actualizar
                if empresa:
                    # Actualizar existente
                    empresa.cliente = cliente
                    empresa.compania = compania
                    if logo_field:
                        empresa.logo = logo_field
                    empresa.save()
                    
                    updated += 1
                    self.stdout.write(self.style.WARNING(f"↻ ACTUALIZADA: {compania}"))
                else:
                    # Crear nueva
                    empresa = Empresa.objects.create(
                        cliente=cliente,
                        compania=compania,
                        codigo=codigo,
                        telefono='',
                        correo='',
                        pais='',
                        logo=logo_field or ''
                    )
                    
                    created += 1
                    self.stdout.write(self.style.SUCCESS(f"✓ CREADA: {compania}"))

            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"✗ ERROR: {e}"))

        # Resumen final
        self.stdout.write("\n" + "=" * 70)
        self.stdout.write(self.style.SUCCESS(f"✅ IMPORTACIÓN COMPLETADA"))
        self.stdout.write("=" * 70)
        self.stdout.write(f"  📊 Total procesadas: {total_rows - 1}")
        self.stdout.write(self.style.SUCCESS(f"  ✓ Creadas:          {created}"))
        self.stdout.write(self.style.WARNING(f"  ↻ Actualizadas:     {updated}"))
        self.stdout.write(f"  ⊘ Saltadas:         {skipped}")
        if errors:
            self.stdout.write(self.style.ERROR(f"  ✗ Errores:          {errors}"))
        self.stdout.write("=" * 70)